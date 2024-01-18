from tkinter import *
import tkinter as tk
from tkinter import font,ttk,Label
import os
from Support_Function import *
from Personal_GUI import PersonalGUI
import datetime
import sys
import pandas as pd
import threading
from openpyxl import load_workbook
# Path Config
USS_CENTRALIZATION_PATH         = r'\\bosch.com\dfsRB\DfsVN\LOC\Hc\RBVH\20_EDA\10_EDA9\01_Internal\10_Tool\00_ToolChain_Template'
    # Path Logo
LOG_PATH                        = os.path.join(USS_CENTRALIZATION_PATH, 'Log')
LOGO_PATH                       = os.path.join(USS_CENTRALIZATION_PATH, 'Icon')

# DBC_GENERATION_PATH             = os.path.join(USS_CENTRALIZATION_PATH,'1.DBC_Generation')

# IMPORT_DBC_PATH                 = os.path.join(USS_CENTRALIZATION_PATH,'2.Import_DBC')
# IMPORT_DBC_FOLDER_PATH          = os.path.join(IMPORT_DBC_PATH, 'Import DBC')
# IMPORT_DBC_LINK_PATH            = os.path.join(IMPORT_DBC_PATH,'README.txt')

# VSM_GENERATION_PATH             = os.path.join(USS_CENTRALIZATION_PATH,'3.VSM_Generation')
# VSM_GENERATION_LINK_PATH        = os.path.join(VSM_GENERATION_PATH,'README.txt')

# TX_OFFSETTUNNING_PATH           = os.path.join(USS_CENTRALIZATION_PATH,'4.Tx_OffsetTunning')
# TX_OFFSETTUNNING_LINK_PATH      = os.path.join(TX_OFFSETTUNNING_PATH,'README.txt')

# LAB_MONITORING_PATH             = os.path.join(USS_CENTRALIZATION_PATH,'5.Lab_Remote')
# VSM_CONVERSION_ASW_PATH         = os.path.join(USS_CENTRALIZATION_PATH,'6.VSM_ConversionASW')
# OPL_TEMPLATE_PATH               = os.path.join(USS_CENTRALIZATION_PATH,'7.OPL_Template')
# REVIEW_CHECKLIST_PATH           = os.path.join(USS_CENTRALIZATION_PATH,'8.Review_Checklist')
# FH_DESIGN_TEMPLATE_PATH         = os.path.join(USS_CENTRALIZATION_PATH,'9.FH_Template')
# RESOURCE_MEASUREMENT_PATH       = os.path.join(USS_CENTRALIZATION_PATH,'10.ResourceManagement_Template')
# DIAG_COMMON_UNDERSTANDING_PATH  = os.path.join(USS_CENTRALIZATION_PATH,'11.DiagCommonUnderstanding')
# FIM_DESIGN_TEMPLATE_PATH        = os.path.join(USS_CENTRALIZATION_PATH,'12.FiM_Design_Template')

# RUNTIME_CHECKLIST_PATH          = os.path.join(USS_CENTRALIZATION_PATH,'13.Runtime_Checklist')
# RUNTIME_CHECKLIST_LINK_PATH     = os.path.join(RUNTIME_CHECKLIST_PATH,'README.txt')

# DOORS_UPDATE_PATH               = os.path.join(USS_CENTRALIZATION_PATH,'14.Doors_update_requirement_linking')

# FAQ_PATH                        = os.path.join(USS_CENTRALIZATION_PATH,'FAQ')
# FAQ_LINK_PATH                   = os.path.join(FAQ_PATH,'README.txt')

TOOLS_VERSION_PATH              = os.path.join(USS_CENTRALIZATION_PATH,'Tools_Version')
USS_KIT_VERSION_PATH            = os.path.join(TOOLS_VERSION_PATH,'USS_KIT_Version.txt')

USS_KIT                         = os.path.join(USS_CENTRALIZATION_PATH,'USS_KIT')

CONFIG_PATH                     = os.path.join(USS_CENTRALIZATION_PATH,'Config')
BUTTON_CONFIG_XLSX_PATH         = os.path.join(CONFIG_PATH,'Button_Config.xlsx')
STATE_CONFIG_PATH               = os.path.join(CONFIG_PATH,'State.txt')
SW_VER = 'SW_1.0'
# Function hover
def on_enter(e):
    # Get current color
    bg = e.widget['background']
    # Convert to RGB
    r, g, b = root.winfo_rgb(bg)
    # Improve brightness, rgb 16 bit
    r, g, b = min(r * 1.5, 65535), min(g * 1.5, 65535), min(b * 1.5, 65535)
    # Scale 0-255,rgb 8 bit
    e.widget['background'] = '#%02x%02x%02x' % (int(r // 256), int(g // 256), int(b // 256))

# Function not hover
def on_leave(e):
    e.widget['background'] = e.widget.default_bg 

#Function threading
def check_update_thread():
    check_version(SW_VER,USS_KIT_VERSION_PATH,USS_KIT)
#     check_update(LAB_MONITORING_PATH, 'LabMonitoringTool.exe',progress_bar,message_label)
    
# Dictionary of buttons and corresponding folder paths
df_paths = pd.read_excel(BUTTON_CONFIG_XLSX_PATH, sheet_name='Open_Folder')
button_paths = df_paths.set_index('button_name')['folder_path'].to_dict()
# button_paths = {
#     'FH Design template'                : FH_DESIGN_TEMPLATE_PATH,
#     'FiM Design template'               : FIM_DESIGN_TEMPLATE_PATH,
#     'OPL Template'                      : OPL_TEMPLATE_PATH,
#     'Resource measurement template'     : RESOURCE_MEASUREMENT_PATH,
#     'Diag - common understanding'       : DIAG_COMMON_UNDERSTANDING_PATH,
#     'Doors update, requirement, linking': DOORS_UPDATE_PATH,
#     'VSM Conversion ASW'                : VSM_CONVERSION_ASW_PATH,
#     'Review checklist'                  : REVIEW_CHECKLIST_PATH,
#     'DBC Generation'                    : DBC_GENERATION_PATH,
#     'Lab Monitoring'                    : LAB_MONITORING_PATH,
#     # Add more if needed
# }

# Dictionary of buttons and corresponding links
df_links = pd.read_excel(BUTTON_CONFIG_XLSX_PATH, sheet_name='Open_URL')
button_links = df_links.set_index('button_name')['link_path'].to_dict()
# button_links = {
#     'Runtime checklist'                 : RUNTIME_CHECKLIST_LINK_PATH,
#     'Tx offset tuning'                  : TX_OFFSETTUNNING_LINK_PATH,
#     'FAQ'                               : FAQ_LINK_PATH,
#     # Add more if needed
# }

# Dictionary of buttons and corresponding exe file
button_exe = {
    # 'Lab Monitoring'                    : {'path': LAB_MONITORING_PATH, 'exe': 'LabMonitoringTool.exe'}
    # Add more if needed
}

# Dictionary of buttons and corresponding folder and link
df_folder_and_link = pd.read_excel(BUTTON_CONFIG_XLSX_PATH, sheet_name='Open_Folder_URL')
button_folder_and_link = df_folder_and_link.set_index('button_name').T.to_dict()
# button_folder_and_link = {
#     'Import DBC'                        : {'folder_path': IMPORT_DBC_FOLDER_PATH, 'link_path': IMPORT_DBC_LINK_PATH},
#     'VSM Generator'                     : {'folder_path': VSM_GENERATION_PATH, 'link_path': VSM_GENERATION_LINK_PATH}
#     # Add more if needed
# }

# Get system log
sys.stdout = open(os.path.join(LOG_PATH,'sys.txt'), 'a')
sys.stderr = sys.stdout
print('================================================================================================================================================================')
print('Time:', datetime.datetime.now())
print('User name:   '+get_myComputer())

# Main GUI
root = Tk()
root.title('USS KIT')
root.grid_columnconfigure(0,weight=1)
root.resizable(False, False)
max_row=0
max_column=0
with open(STATE_CONFIG_PATH, 'r') as f:
    states = f.read().splitlines()
    sheet_var = StringVar()
for state in states:
    df = pd.read_excel(BUTTON_CONFIG_XLSX_PATH, sheet_name=state)
    df = df.dropna(how='all')
    max_row = max(max_row, int(df['row'].max()))
    max_column = max(max_column, int(df['column'].max()))
df = pd.read_excel(BUTTON_CONFIG_XLSX_PATH,sheet_name='Dev_Tools')
df = df.dropna(how='all')
sheet_var = StringVar(value='Dev_Tools')
sheet_combobox = ttk.Combobox(root, textvariable=sheet_var,state='readonly',width=9)
sheet_combobox['values'] = states
sheet_combobox.grid(row=max_row+2, column=0)

try:
    photo = PhotoImage(file = os.path.join(LOGO_PATH,'logo.png'))
    root.iconphoto(False, photo)
except:
    pass
root.configure(bg='white')

title = Label(root, text="USS KIT", bg="#0d6759", fg="white", font = font.Font(size=13, weight='bold'),height= 2, relief='raised', borderwidth=5)
title.grid(row=0, column=0, columnspan=max_column+1,sticky='ew')

canvas = Canvas(root, width=60,bg="#0d6759",highlightthickness=0,relief='raised', borderwidth=5)
canvas.grid(row=1, column=0, rowspan=max_row, sticky='nsew')
text = canvas.create_text(45,210,text="Tools, template, checklist, design", fill="white", font=font.Font(size=13, weight='bold'))
# Rotate text direction to vertical
canvas.itemconfig(text, angle=90)

def get_colors_from_excel():
    # Load the workbook and select the "Colors" sheet
    df = pd.read_excel(BUTTON_CONFIG_XLSX_PATH,sheet_name='Colors')
    
    # Get the colors from the "Background" and "Text fill" columns
    background_colors = df['Background'].dropna().tolist()
    text_fill_colors = df['Text fill'].dropna().tolist()
    return background_colors, text_fill_colors

#Add buttons
def open_new_window():
    new_window = Toplevel(root)
    x = root.winfo_x()
    y = root.winfo_y()

    new_window.geometry("+%d+%d" % (x + 100, y + 100))  # Position the new window 100px to the right and 100px below the main window
    new_window.attributes('-topmost', True)
    # Create entry fields for text, row, column, background and text fill
    fields = ['Text', 'Row', 'Column', 'Background', 'Text fill']
    entries = {field: Entry(new_window, width=20) for field in fields}
    bg,fg = get_colors_from_excel()
    # Create comboboxes for background and text fill colors
    color_comboboxes = {
        'Background': ttk.Combobox(new_window, values=bg, width=17,state='readonly'),
        'Text fill': ttk.Combobox(new_window, values=fg, width=17,state='readonly')
    }
    sheet_combobox = ttk.Combobox(new_window, values=states, width=17, state='readonly')
    entries['Usage'] = sheet_combobox

    # Position the entry fields and comboboxes in the new window
    for i, field in enumerate(fields+['Usage']):
        Label(new_window, text=field).grid(row=i, column=0, sticky='w',padx=5,pady=5)
        if field in ['Background', 'Text fill']:
            color_comboboxes[field].grid(row=i, column=1, sticky='w',padx=5,pady=5)
        elif field == 'Usage':
            sheet_combobox.grid(row=i, column=1, sticky='w',padx=5,pady=5)
        else:
            entries[field].grid(row=i, column=1, sticky='w',padx=5,pady=5)

    # Create the "Add button" button in the new window
    add_button = Button(new_window, text='Add button', command=lambda: check_password(entries, color_comboboxes,new_window))
    add_button.grid(row=len(fields)+1, column=0, columnspan=2,pady=10)

def check_password(entries, color_comboboxes,new_window):
    # Create a new window for password entry
    password_window = Toplevel(root)
    x = new_window.winfo_rootx()
    y = new_window.winfo_rooty()
    width = new_window.winfo_width()
    height = new_window.winfo_height()
    password_window.geometry("+%d+%d" % (x + width, y + height - 100))
    password_window.attributes('-topmost', True)
    # Create a label and entry field for the password
    Label(password_window, text="Password").grid(row=0, column=0)
    password_entry = Entry(password_window, show="*")
    password_entry.grid(row=0, column=1)

    # Create a "Submit" button that checks the password
    submit_button = Button(password_window, text="Submit", command=lambda: validate_password(entries, color_comboboxes, password_entry, password_window,new_window))
    submit_button.grid(row=1, column=0, columnspan=2)

def validate_password(entries, color_comboboxes, password_entry, password_window,new_window):
    # Check if the entered password is correct
    if password_entry.get() == "usskit":
        # If the password is correct, close the password window and add to excel
        password_window.destroy()
        add_to_excel(entries, color_comboboxes)
        new_window.destroy()
    else:
        # If the password is incorrect, show an error message
        password_window.attributes('-topmost', False)
        messagebox.showerror("Error", "Incorrect password. Please try again.")
        password_entry.delete(0, 'end')
        password_window.attributes('-topmost', True)
def add_to_excel(entries, color_comboboxes):
    # Load the workbook and select the "Position_Color" sheet
    wb = load_workbook(BUTTON_CONFIG_XLSX_PATH)
    ws = wb[entries['Usage'].get()]

    # Find the first empty row
    first_empty_row = next((i for i, row in enumerate(ws.iter_rows(values_only=True)) if all(cell is None for cell in row)), None)

    # If there is no empty row, append to the last row
    if first_empty_row is None:
        first_empty_row = ws.max_row

    # Write the values from the entry fields and comboboxes to the first empty row
    ws.cell(row=first_empty_row+1, column=1, value=str(entries['Text'].get()))
    ws.cell(row=first_empty_row+1, column=2, value=int(entries['Row'].get()))
    ws.cell(row=first_empty_row+1, column=3, value=int(entries['Column'].get()))
    ws.cell(row=first_empty_row+1, column=4, value=str(color_comboboxes['Background'].get()))
    ws.cell(row=first_empty_row+1, column=5, value=str(color_comboboxes['Text fill'].get()))

    # Save the workbook
    wb.save(BUTTON_CONFIG_XLSX_PATH)
    
    messagebox.showinfo("Success", "Button has been successfully added, please reopen the application to see the update!")

add_button = Button(root,text= 'Add button',relief='raised', borderwidth=5,bg='#ffd166',activebackground='#ffd166', fg='#000',width=10)
add_button.default_bg = '#ffd166'
add_button.bind("<Enter>", on_enter) 
add_button.bind("<Leave>", on_leave) 
add_button['command'] = open_new_window
add_button.grid(row=max_row+1,column=0)
Persional = Button(root,text= 'Persional',relief='raised', borderwidth=5,bg='#ffd166',activebackground='#ffd166', fg='#000',width=10)
Persional.default_bg = '#ffd166'
Persional.bind("<Enter>", on_enter) 
Persional.bind("<Leave>", on_leave) 
Persional['command'] = PersonalGUI
Persional.grid(row=max_row,column=0)
# buttons = df.to_dict('records')

# buttons = [
#     # Row 1
#     {'text': 'DBC Generation', 'row': 1, 'column': 1, 'bg': '#90e0ef', 'fg': '#03045e'},
#     {'text': 'Import DBC', 'row': 1, 'column': 2, 'bg': '#90e0ef', 'fg': '#03045e'},
#     {'text': 'VSM Generator', 'row': 1, 'column': 3, 'bg': '#90e0ef', 'fg': '#03045e'},
#     {'text': 'Tx offset tuning', 'row': 1, 'column': 4, 'bg': '#90e0ef', 'fg': '#03045e'},
#     {'text': 'Lab Monitoring', 'row': 1, 'column': 5, 'bg': '#90e0ef', 'fg': '#03045e'},
    
#     # Row 2
#     {'text': 'Diag - common understanding', 'row': 2, 'column': 1, 'bg': '#00b4d8', 'fg': '#03045e'},
#     {'text': 'FH Design template', 'row': 2, 'column': 2, 'bg': '#00b4d8', 'fg': '#03045e'},
#     {'text': 'FiM Design template', 'row': 2, 'column': 3, 'bg': '#00b4d8', 'fg': '#03045e'},
#     {'text': 'Doors update, requirement, linking', 'row': 2, 'column': 4, 'bg': '#00b4d8', 'fg': '#03045e'},
#     {'text': 'VSM Conversion ASW', 'row': 2, 'column': 5, 'bg': '#00b4d8', 'fg': '#03045e'},
    
#     # Row 3
#     {'text': 'OPL Template', 'row': 3, 'column': 1, 'bg': '#0077b6', 'fg': '#03045e'},
#     {'text': 'Review checklist', 'row': 3, 'column': 2, 'bg': '#0077b6', 'fg': '#03045e'},
#     {'text': 'Runtime checklist', 'row': 3, 'column': 3, 'bg': '#0077b6', 'fg': '#03045e'},
#     {'text': 'Resource measurement template', 'row': 3, 'column': 4, 'bg': '#0077b6', 'fg': '#03045e'},
#     {'text': 'FAQ', 'row': 3, 'column': 5, 'bg': '#0077b6', 'fg': '#03045e'}
#     # Add more if needed
# ]
progress_bar = ttk.Progressbar(root, length=350, mode='indeterminate')
progress_bar.grid(row=max_row+3, column=1, columnspan=6,padx = 5,sticky='e')
message_label = Label(root, text="", bg="white", fg="black", font=font.Font(size=12))
message_label.grid(row=max_row+3, column=0, columnspan=6, sticky='w')
# Thread to check update when open GUI
update_thread = threading.Thread(target=check_update_thread)
update_thread.start()
# create a canvas
canvas = Canvas(root,width= (5)*203, height=(3)*138)
canvas.grid(row=1, column=1)

# create vertical scrollbar
yscrollbar = Scrollbar(root, orient='vertical', command=canvas.yview)
yscrollbar.grid(row=1, column=max_column+1, sticky='ns')

# create horizontal scrollbar
xscrollbar = Scrollbar(root, orient='horizontal', command=canvas.xview)
xscrollbar.grid(row=max_row+1, column=1, sticky='ew')

# connect scrollbars with canvas
canvas.configure(yscrollcommand=yscrollbar.set, xscrollcommand=xscrollbar.set)

# create frame combine button
frame = Frame(canvas)

# add frame to canvas
canvas.create_window((0, 0), window=frame, anchor='nw')
def update_buttons():
    for widget in frame.winfo_children():
            widget.destroy()
    df = pd.read_excel(BUTTON_CONFIG_XLSX_PATH, sheet_name=sheet_var.get())
    df = df.dropna(how='all')
    buttons = df.to_dict('records')

    for button in buttons:
        b = Button(frame, text=button['text'], bg=button['bg'],activebackground=button['bg'], fg=button['fg'],width = 17, height= 5,wraplength=150,font=font.Font(size=12), relief='raised', borderwidth=5)
        b.default_bg = button['bg']  
        b.bind("<Enter>", on_enter) 
        b.bind("<Leave>", on_leave) 
        b.grid(row=button['row'], column=button['column'])
        if button['text'] in button_links:
            link_path = button_links[button['text']]
            b.config(command=lambda link_path=link_path, progress_bar=progress_bar: threading.Thread(target=open_link, args=(link_path, progress_bar,message_label)).start())   
        elif button['text'] in button_paths:
            folder_path = button_paths[button['text']]
            b.config(command=lambda folder_path=folder_path, progress_bar=progress_bar: threading.Thread(target=open_folder, args=(folder_path, progress_bar,message_label)).start())   
        elif button['text'] in button_exe:
            exe_info = button_exe[button['text']]
            b.config(command=lambda exe_info=exe_info,progress_bar=progress_bar: threading.Thread(target=open_file, args=(exe_info['path'],exe_info['exe'], progress_bar,message_label)).start())
        elif button['text'] in button_folder_and_link:
            folder_and_link_info = button_folder_and_link[button['text']]
        
            b.config(command=lambda folder_and_link_info=folder_and_link_info, progress_bar=progress_bar: threading.Thread(target=open_folder_and_link, args=(folder_and_link_info['folder_path'], folder_and_link_info['link_path'], progress_bar, message_label)).start())
    # canvas.config(scrollregion=canvas.bbox('all'))
    root.after_idle(lambda: canvas.config(scrollregion=canvas.bbox('all')))
root.update()
sheet_combobox.bind("<<ComboboxSelected>>", lambda _: update_buttons())
update_buttons()
# canvas.config(scrollregion=canvas.bbox('all'))
root.mainloop()
sys.stdout.close()
